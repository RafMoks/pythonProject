import os
import re
from requests import get
import openpyxl  # импортируем библиотеку для работы с файлами эксель
from openpyxl.styles import Font, PatternFill
import calendar
import shutil


def podgotovka(date_year, date_month):
    name = f'{date_month}.{date_year}.xlsx'
    shutil.copyfile('Шаблон.xlsx', name)
    x = os.getcwd()
    x2 = os.listdir(x)
    if 'ELIRA' not in x2:
        os.mkdir(f'{x}\ELIRA\\')

    name = shutil.move(name, fr'{x}\ELIRA\\')
    return name
def get_data(url):
    flag, flag2 = False, False
    d = {}
    r = get(url).text
    for c in r.split('\n'):
        if re.match(
                r'<th class="center aligned widthconst2" colspan="4">Ежедневный прогноз часов пиковой нагрузки</th>',
                c.strip()) or flag:
            flag = True
            data = re.search(r'(<strong>)(\d\d.\d\d.\d\d\d\d)(</strong>)', c.strip())
            time_chas = re.search(r'(<span>)(\d\d)(</span>)', c.strip())
            if data and not flag2:
                data2 = data.group(2)
                d[data2] = []
                flag2 = True
            elif time_chas is not None:
                if len(d[data2]) < 2:
                    d[data2].append(time_chas.group(2))
                else:
                    d[data2].append(time_chas.group(2))
                    flag2 = False
    return d
def perenos_data(dikt, date_year, date_month, put):
    # открытие файла с показаниями
    wb = openpyxl.load_workbook(filename=put)
    sheet = wb['Лист1']  # делаем активный лист
    nb_row = sheet.max_row  # находим максимальное количество строк для цикла
    nb_col = sheet.max_column  # находим максимальное количество столбцов для цикла
    month_t = month_table(date_year, date_month)
    for i in range(len(month_t)):
        sheet.cell(row=10, column=i+2, value=month_t[i])
    for i in range(2, nb_col + 1):
        x2 = sheet.cell(row=10, column=i).value
        if x2 in dikt:
            for j in range(10, nb_row + 1):
                x = sheet.cell(row=j, column=i).value
                if str(x) in dikt[x2]:
                    stil = dikt[x2].index(str(x))
                    if stil == 0:
                        stil2 = 'FF0000'
                    elif stil == 1:
                        stil2 = 'FF8C00'
                    else:
                        stil2 = 'FFFF00'
                    d = sheet.cell(row=j, column=i, value=x)
                    d.fill = PatternFill('solid', fgColor=stil2)  # меняем цвет

    wb.save(put)  # сохраняем изменения в файле
    wb.close()
    month_table(date_year, date_month)
def month_table(date_year, date_month):
    return [f'{"0"+str(j) if j < 10 else j}.{date_month}.{date_year}' for c in calendar.monthcalendar(int(date_year), int(date_month)) for j in c if j != 0]


def main():
    print('Привет, я программа для парсинга с сайта elira.pro\n'
          'для моей работы необходим интернет, а так же желаемые год и месяц.')
    date_year = input('Введите корректный год в формате ГГГГ  ')
    date_month = input('Введите корректный месяц в формате ММ  ')
    try:
        put = podgotovka(date_year, date_month)
        dikt = get_data(f"http://elira.pro/Forecast/PeakhourContent/29/{date_year}/{date_month}")
        perenos_data(dikt, date_year, date_month, put)
    except Exception as err:
        print(err)
    input('Нажмите Enter')

if __name__ == '__main__':
    main()